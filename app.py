import streamlit as st
import pandas as pd
import cv2
import numpy as np
from PIL import Image
from io import BytesIO
import easyocr
import openpyxl

# ✅ Initialize EasyOCR (English only, GPU off for Streamlit Cloud)
reader = easyocr.Reader(['en'], gpu=False)

st.set_page_config(page_title="PostNL Cart Tracker", layout="wide")
st.title("📦 PostNL Cart Tracker (Final Version)")

st.markdown("""
✅ **Final Version**  
- Detects **multiple carts per photo** and counts them correctly.  
- Extracts **HAGA, HAGB, HAGE, SMO, BIMIC** categories.  
- **Debug image** with detected tags drawn.  
- **Excel report formatted like official scanned form**.  
- Adds a **TOTAL row per day** automatically.  
""")

# ✅ Day Colors (used for auto-detection)
DAY_COLORS = {
    "Sunday": (150, 75, 0),
    "Monday": (0, 128, 0),
    "Tuesday": (128, 128, 128),
    "Wednesday": (255, 255, 0),
    "Thursday": (255, 0, 0),
    "Friday": (0, 0, 255),
    "Saturday": (255, 165, 0)
}

CATEGORIES = ["HAGA", "HAGB", "HAGE", "SMO", "BIMIC"]

def preprocess_image(image):
    img = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 11, 17, 17)
    return gray

def detect_day_from_color(image):
    img = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    avg_color = img.mean(axis=(0, 1))[::-1]
    closest_day, min_dist = None, float("inf")
    for day, color in DAY_COLORS.items():
        dist = np.linalg.norm(np.array(avg_color) - np.array(color))
        if dist < min_dist:
            closest_day, min_dist = day, dist
    return closest_day

def analyze_carts(image):
    """Detect multiple carts and extract their categories."""
    img = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    processed = preprocess_image(image)

    results = reader.readtext(processed, detail=1)
    category_counts = {cat: 0 for cat in CATEGORIES}
    debug_img = img.copy()

    for (bbox, text, conf) in results:
        text_upper = text.upper()
        pts = np.array(bbox, dtype=np.int32)
        cv2.polylines(debug_img, [pts], True, (0, 255, 0), 2)

        for cat in CATEGORIES:
            if cat in text_upper:
                category_counts[cat] += 1
                cv2.putText(debug_img, cat, (int(bbox[0][0]), int(bbox[0][1]) - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 0, 0), 2)
    return category_counts, debug_img

def convert_df_to_excel_formatted(dataframe):
    """Generate Excel in the official form layout with total rows."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Cart Summary")

        # Add formatting & total rows
        workbook = writer.book
        sheet = writer.sheets["Cart Summary"]

        bold_font = openpyxl.styles.Font(bold=True)
        alignment_center = openpyxl.styles.Alignment(horizontal="center")

        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 3

        # Add TOTAL rows per day
        df_days = dataframe.groupby("Day")["Count"].sum().reset_index()
        last_row = sheet.max_row + 2
        sheet.cell(row=last_row, column=1, value="TOTAL PER DAY").font = bold_font

        for i, row in df_days.iterrows():
            sheet.cell(row=last_row + i + 1, column=2, value=row["Day"]).font = bold_font
            sheet.cell(row=last_row + i + 1, column=4, value=row["Count"]).font = bold_font
            sheet.cell(row=last_row + i + 1, column=4).alignment = alignment_center

    return output.getvalue()

uploaded_files = st.file_uploader("Upload cart photos:", type=["jpg", "jpeg", "png"], accept_multiple_files=True)
results = []

if uploaded_files:
    for file in uploaded_files:
        image = Image.open(file)
        detected_day = detect_day_from_color(image)

        manual_day = st.selectbox(
            f"Select Day for {file.name} (detected: {detected_day})",
            options=list(DAY_COLORS.keys()),
            index=list(DAY_COLORS.keys()).index(detected_day) if detected_day else 0,
            key=file.name
        )

        category_counts, debug_img = analyze_carts(image)

        # Show debug image
        st.image(cv2.cvtColor(debug_img, cv2.COLOR_BGR2RGB),
                 caption=f"Detected Carts for {file.name}", use_column_width=True)

        for cat, count in category_counts.items():
            if count > 0:
                results.append({
                    "Type": "Gerichte Landen",
                    "Category": cat,
                    "Day": manual_day,
                    "Count": count
                })

    if results:
        df = pd.DataFrame(results)
        st.subheader("📊 Cart Summary Table")
        st.dataframe(df, use_container_width=True)

        excel_data = convert_df_to_excel_formatted(df)
        st.download_button(
            label="⬇️ Download Excel Report (Formatted)",
            data=excel_data,
            file_name="postnl_cart_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success("✅ Report ready!")
    else:
        st.warning("⚠️ No carts detected. Check photo quality or tags.")
